import tkinter as tk
from tkinter import ttk, messagebox
import os
import sys
import webbrowser

# Добавляем родительскую директорию в путь поиска модулей
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from database import Database


class HotelListPanel(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent)

        self.parent = parent
        self.db = Database()

        # Получаем список отелей из базы данных
        self.hotels = []

        # Текущий выбранный отель и его обзоры
        self.selected_hotel = None
        self.hotel_reviews = []

        # Создаем интерфейс
        self.create_widgets()

        # Загружаем список отелей
        self.refresh_hotel_list()

    def create_widgets(self):
        """Создание виджетов панели списка отелей."""
        # Верхняя часть - заголовок и поиск
        self.header_frame = ttk.Frame(self)
        self.header_frame.pack(fill=tk.X, pady=(0, 10))

        self.title_label = ttk.Label(self.header_frame, text="Список отелей",
                                     style="Title.TLabel")
        self.title_label.pack(pady=10)

        # Панель поиска
        self.search_frame = ttk.Frame(self.header_frame)
        self.search_frame.pack(fill=tk.X, padx=20, pady=5)

        self.search_label = ttk.Label(self.search_frame, text="Поиск:")
        self.search_label.pack(side=tk.LEFT, padx=(0, 5))

        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(self.search_frame, textvariable=self.search_var, width=40)
        self.search_entry.pack(side=tk.LEFT, padx=5)

        self.search_button = ttk.Button(self.search_frame, text="Найти",
                                        command=self.search_hotels)
        self.search_button.pack(side=tk.LEFT, padx=5)

        self.reset_button = ttk.Button(self.search_frame, text="Сбросить",
                                       command=self.reset_search)
        self.reset_button.pack(side=tk.LEFT, padx=5)

        # Разделение на список отелей и детали отеля
        self.content_frame = ttk.Frame(self)
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=10)

        # Панель со списком отелей (левая часть)
        self.hotels_frame = ttk.LabelFrame(self.content_frame, text="Отели")
        self.hotels_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

        # Создаем таблицу отелей
        self.create_hotel_table()

        # Панель с деталями отеля (правая часть)
        self.details_frame = ttk.LabelFrame(self.content_frame, text="Информация об отеле")
        self.details_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))

        # Создаем панель деталей отеля
        self.create_hotel_details()

        # Нижняя часть - кнопки действий
        self.buttons_frame = ttk.Frame(self)
        self.buttons_frame.pack(fill=tk.X, pady=10, padx=10)

        self.refresh_button = ttk.Button(self.buttons_frame, text="Обновить список",
                                         command=self.refresh_hotel_list)
        self.refresh_button.pack(side=tk.LEFT, padx=5)

        self.export_button = ttk.Button(self.buttons_frame, text="Экспорт списка",
                                        command=self.export_hotel_list)
        self.export_button.pack(side=tk.RIGHT, padx=5)

    def create_hotel_table(self):
        """Создание таблицы отелей."""
        # Создаем фрейм для таблицы
        self.table_frame = ttk.Frame(self.hotels_frame)
        self.table_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Создаем Treeview для отображения списка отелей
        columns = ('id', 'name', 'rating', 'reviews')
        self.hotel_table = ttk.Treeview(self.table_frame, columns=columns, show='headings')

        # Настраиваем заголовки столбцов
        self.hotel_table.heading('id', text='ID')
        self.hotel_table.heading('name', text='Название отеля')
        self.hotel_table.heading('rating', text='Рейтинг')
        self.hotel_table.heading('reviews', text='Кол-во отзывов')

        # Настраиваем ширину столбцов
        self.hotel_table.column('id', width=50, anchor=tk.CENTER)
        self.hotel_table.column('name', width=250)
        self.hotel_table.column('rating', width=80, anchor=tk.CENTER)
        self.hotel_table.column('reviews', width=100, anchor=tk.CENTER)

        # Добавляем полосу прокрутки
        scrollbar = ttk.Scrollbar(self.table_frame, orient=tk.VERTICAL, command=self.hotel_table.yview)
        self.hotel_table.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.hotel_table.pack(fill=tk.BOTH, expand=True)

        # Обработчик выбора отеля
        self.hotel_table.bind('<<TreeviewSelect>>', self.on_hotel_select)

    def create_hotel_details(self):
        """Создание панели деталей отеля."""
        # Создаем фрейм для деталей отеля
        self.details_content = ttk.Frame(self.details_frame)
        self.details_content.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Заголовок с названием отеля
        self.hotel_name_var = tk.StringVar(value="Выберите отель из списка")
        self.hotel_name_label = ttk.Label(self.details_content,
                                          textvariable=self.hotel_name_var,
                                          font=("Arial", 14, "bold"),
                                          wraplength=300)
        self.hotel_name_label.pack(pady=(0, 10))

        # Рейтинг отеля (звезды)
        self.rating_frame = ttk.Frame(self.details_content)
        self.rating_frame.pack(fill=tk.X, pady=(0, 10))

        self.rating_label = ttk.Label(self.rating_frame, text="Рейтинг: ")
        self.rating_label.pack(side=tk.LEFT)

        self.stars_frame = ttk.Frame(self.rating_frame)
        self.stars_frame.pack(side=tk.LEFT)

        # Адрес отеля
        self.address_var = tk.StringVar()
        self.address_frame = ttk.Frame(self.details_content)
        self.address_frame.pack(fill=tk.X, pady=(0, 10))

        self.address_label = ttk.Label(self.address_frame, text="Адрес: ")
        self.address_label.pack(side=tk.LEFT, anchor=tk.N)

        self.address_value = ttk.Label(self.address_frame,
                                       textvariable=self.address_var,
                                       wraplength=300)
        self.address_value.pack(side=tk.LEFT)

        # Notebook для отзывов
        self.details_notebook = ttk.Notebook(self.details_content)
        self.details_notebook.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        # Вкладка с отзывами
        self.reviews_frame = ttk.Frame(self.details_notebook)
        self.details_notebook.add(self.reviews_frame, text="Отзывы")

        # Настройка вкладки с отзывами
        self.reviews_canvas = tk.Canvas(self.reviews_frame, highlightthickness=0)
        self.reviews_scrollbar = ttk.Scrollbar(self.reviews_frame,
                                               orient="vertical",
                                               command=self.reviews_canvas.yview)
        self.reviews_scrollframe = ttk.Frame(self.reviews_canvas)

        self.reviews_scrollframe.bind(
            "<Configure>",
            lambda e: self.reviews_canvas.configure(
                scrollregion=self.reviews_canvas.bbox("all")
            )
        )

        self.reviews_canvas.create_window((0, 0), window=self.reviews_scrollframe, anchor="nw")
        self.reviews_canvas.configure(yscrollcommand=self.reviews_scrollbar.set)

        self.reviews_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.reviews_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Сообщение о выборе отеля
        self.no_hotel_message = ttk.Label(self.reviews_scrollframe,
                                          text="Выберите отель из списка для просмотра отзывов",
                                          font=("Arial", 10, "italic"))
        self.no_hotel_message.pack(pady=20)

    def refresh_hotel_list(self):
        """Обновление списка отелей из базы данных."""
        try:
            # Получаем список отелей
            self.hotels = self.db.get_hotels()

            # Очищаем таблицу
            for item in self.hotel_table.get_children():
                self.hotel_table.delete(item)

            # Заполняем таблицу данными
            for hotel in self.hotels:
                hotel_id, name, address, review_count, avg_rating = hotel

                # Форматируем рейтинг как количество звезд
                rating_text = f"{avg_rating}★" if avg_rating else "-"

                self.hotel_table.insert('', 'end', values=(
                    hotel_id, name, rating_text, review_count or 0
                ))

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить список отелей: {e}")

    def on_hotel_select(self, event):
        """Обработчик выбора отеля в таблице."""
        # Получаем ID выбранного отеля
        selected_items = self.hotel_table.selection()

        if not selected_items:
            return

        # Получаем данные выбранного отеля
        item = selected_items[0]
        hotel_id = self.hotel_table.item(item, "values")[0]

        # Загружаем информацию об отеле
        self.load_hotel_details(hotel_id)

    def load_hotel_details(self, hotel_id):
        """Загрузка подробной информации об отеле."""
        try:
            # Получаем информацию из базы данных
            hotel, reviews = self.db.get_hotel_details(int(hotel_id))

            if not hotel:
                messagebox.showerror("Ошибка", "Отель не найден в базе данных.")
                return

            # Сохраняем текущий отель
            self.selected_hotel = hotel
            self.hotel_reviews = reviews

            # Обновляем информацию в интерфейсе
            self.update_hotel_info(hotel, reviews)

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить информацию об отеле: {e}")

    def update_hotel_info(self, hotel, reviews):
        """Обновление информации об отеле в интерфейсе."""
        # Обновляем название отеля
        self.hotel_name_var.set(hotel[1])  # hotel[1] - название отеля

        # Обновляем адрес
        self.address_var.set(hotel[2] or "Адрес не указан")  # hotel[2] - адрес

        # Обновляем рейтинг (звезды)
        for widget in self.stars_frame.winfo_children():
            widget.destroy()

        avg_rating = hotel[4]  # hotel[4] - средний рейтинг
        if avg_rating:
            star_count = int(avg_rating)

            for i in range(5):
                if i < star_count:
                    star_text = "★"  # Закрашенная звезда
                    star_color = "#FFD700"  # Золотой цвет
                else:
                    star_text = "☆"  # Пустая звезда
                    star_color = "#C0C0C0"  # Серый цвет

                star_label = ttk.Label(self.stars_frame, text=star_text,
                                       font=("Arial", 16), foreground=star_color)
                star_label.pack(side=tk.LEFT)
        else:
            ttk.Label(self.stars_frame, text="Нет оценок",
                      font=("Arial", 10, "italic")).pack(side=tk.LEFT)

        # Обновляем отзывы
        # Очищаем фрейм с отзывами
        for widget in self.reviews_scrollframe.winfo_children():
            widget.destroy()

        if not reviews:
            ttk.Label(self.reviews_scrollframe, text="Нет отзывов",
                      font=("Arial", 10, "italic")).pack(pady=20)
            return

        # Добавляем отзывы
        for i, review_data in enumerate(reviews):
            review = review_data['review']
            criteria = review_data['criteria_ratings']

            review_frame = ttk.Frame(self.reviews_scrollframe)
            review_frame.pack(fill=tk.X, pady=(0, 15), padx=5)

            # Разделяем отзывы линией
            if i > 0:
                separator = ttk.Separator(review_frame, orient='horizontal')
                separator.pack(fill=tk.X, pady=(0, 10))

            # Заголовок отзыва
            header_frame = ttk.Frame(review_frame)
            header_frame.pack(fill=tk.X, pady=(5, 10))

            # Рейтинг отзыва
            rating = review[1]  # review[1] - рейтинг
            rating_label = ttk.Label(header_frame, text=f"Рейтинг: {rating}★",
                                     font=("Arial", 10, "bold"))
            rating_label.pack(side=tk.LEFT)

            # Дата отзыва
            date = review[3]  # review[3] - дата
            if date:
                date_parts = date.split(" ")[0].split("-")
                formatted_date = f"{date_parts[2]}.{date_parts[1]}.{date_parts[0]}"
                date_label = ttk.Label(header_frame, text=f"Дата: {formatted_date}",
                                       font=("Arial", 10))
                date_label.pack(side=tk.RIGHT)

            # Группируем критерии по категориям
            categories = {}
            for criteria_rating in criteria:
                category = criteria_rating[3]  # criteria_rating[3] - название категории

                if category not in categories:
                    categories[category] = []

                categories[category].append(criteria_rating)

            # Отображаем критерии по категориям
            for category, criteria_list in categories.items():
                category_frame = ttk.LabelFrame(review_frame, text=category)
                category_frame.pack(fill=tk.X, pady=(0, 5))

                # Создаем таблицу для критериев
                criteria_table = ttk.Treeview(category_frame, columns=('criteria', 'rating', 'description'),
                                              show='headings', height=len(criteria_list))

                criteria_table.heading('criteria', text='Критерий')
                criteria_table.heading('rating', text='Оценка')
                criteria_table.heading('description', text='Описание')

                criteria_table.column('criteria', width=150)
                criteria_table.column('rating', width=70, anchor=tk.CENTER)
                criteria_table.column('description', width=200)

                for criteria_rating in criteria_list:
                    name = criteria_rating[0]  # criteria_rating[0] - название критерия
                    rating = criteria_rating[1]  # criteria_rating[1] - оценка
                    description = criteria_rating[2]  # criteria_rating[2] - описание

                    criteria_table.insert('', 'end', values=(name, rating, description))

                criteria_table.pack(fill=tk.X, expand=True)

    def search_hotels(self):
        """Поиск отелей в базе данных."""
        search_query = self.search_var.get().strip().lower()

        if not search_query:
            self.refresh_hotel_list()
            return

        try:
            # Очищаем таблицу
            for item in self.hotel_table.get_children():
                self.hotel_table.delete(item)

            # Фильтруем отели по запросу
            filtered_hotels = [
                hotel for hotel in self.hotels
                if search_query in hotel[1].lower()  # hotel[1] - название отеля
            ]

            # Заполняем таблицу отфильтрованными данными
            for hotel in filtered_hotels:
                hotel_id, name, address, review_count, avg_rating = hotel

                # Форматируем рейтинг как количество звезд
                rating_text = f"{avg_rating}★" if avg_rating else "-"

                self.hotel_table.insert('', 'end', values=(
                    hotel_id, name, rating_text, review_count or 0
                ))

            # Обновляем статус
            hotel_count = len(filtered_hotels)
            if hotel_count > 0:
                messagebox.showinfo("Результаты поиска", f"Найдено отелей: {hotel_count}")
            else:
                messagebox.showinfo("Результаты поиска", "Отели не найдены")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при поиске отелей: {e}")

    def reset_search(self):
        """Сброс результатов поиска."""
        self.search_var.set("")
        self.refresh_hotel_list()

    def add_api_hotel_to_db(self, hotel_name, address=""):
        """Добавление отеля в базу данных."""
        try:
            # Добавляем отель в базу данных
            hotel_id = self.db.add_hotel(hotel_name, address)

            messagebox.showinfo("Успешно",
                                f"Отель '{hotel_name}' успешно добавлен в базу данных.")

            # Обновляем список отелей
            self.refresh_hotel_list()

            # Предлагаем перейти к оценке отеля
            if messagebox.askyesno("Оценка отеля",
                                   f"Хотите оценить отель '{hotel_name}' прямо сейчас?"):
                if hasattr(self.parent.master, 'new_rating'):
                    self.parent.master.new_rating()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось добавить отель в базу данных: {e}")

    def export_hotel_list(self):
        """Экспорт списка отелей в файл."""
        from tkinter import filedialog
        import csv

        # Запрашиваем имя файла для сохранения
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить список отелей"
        )

        if not file_path:
            return

        try:
            # Если выбран формат CSV
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)

                    # Записываем заголовки
                    writer.writerow(['ID', 'Название', 'Адрес', 'Количество отзывов', 'Средний рейтинг'])

                    # Записываем данные отелей
                    for hotel in self.hotels:
                        writer.writerow(hotel)

            # Если выбран формат Excel
            elif file_path.lower().endswith('.xlsx'):
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment

                    # Создаем новую книгу и лист
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet.title = "Список отелей"

                    # Записываем заголовки
                    headers = ['ID', 'Название', 'Адрес', 'Количество отзывов', 'Средний рейтинг']
                    for col_num, header in enumerate(headers, 1):
                        cell = sheet.cell(row=1, column=col_num)
                        cell.value = header
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

                    # Записываем данные отелей
                    for row_num, hotel in enumerate(self.hotels, 2):
                        for col_num, value in enumerate(hotel, 1):
                            cell = sheet.cell(row=row_num, column=col_num)
                            cell.value = value

                    # Автоматическая настройка ширины столбцов
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter

                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))

                        adjusted_width = (max_length + 2) * 1.2
                        sheet.column_dimensions[column_letter].width = adjusted_width

                    # Сохраняем книгу
                    workbook.save(file_path)

                except ImportError:
                    messagebox.showwarning("Предупреждение",
                                           "Для экспорта в Excel требуется модуль openpyxl. " +
                                           "Будет создан файл CSV.")

                    # Если модуль не установлен, сохраняем в CSV
                    csv_path = file_path.replace('.xlsx', '.csv')

                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)

                        # Записываем заголовки
                        writer.writerow(['ID', 'Название', 'Адрес', 'Количество отзывов', 'Средний рейтинг'])

                        # Записываем данные отелей
                        for hotel in self.hotels:
                            writer.writerow(hotel)

                    file_path = csv_path

            messagebox.showinfo("Экспорт завершен",
                                f"Список отелей успешно экспортирован в файл:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Ошибка",
                                 f"Не удалось экспортировать список отелей: {e}").BOTH, expand = True, padx = 5, pady = 5)

            # Создаем Treeview для отображения списка отелей
            columns = ('id', 'name', 'rating', 'reviews')
            self.hotel_table = ttk.Treeview(self.table_frame, columns=columns, show='headings')

            # Настраиваем заголовки столбцов
            self.hotel_table.heading('id', text='ID')
            self.hotel_table.heading('name', text='Название отеля')
            self.hotel_table.heading('rating', text='Рейтинг')
            self.hotel_table.heading('reviews', text='Кол-во отзывов')

            # Настраиваем ширину столбцов
            self.hotel_table.column('id', width=50, anchor=tk.CENTER)
            self.hotel_table.column('name', width=250)
            self.hotel_table.column('rating', width=80, anchor=tk.CENTER)
            self.hotel_table.column('reviews', width=100, anchor=tk.CENTER)

            # Добавляем полосу прокрутки
            scrollbar = ttk.Scrollbar(self.table_frame, orient=tk.VERTICAL, command=self.hotel_table.yview)
            self.hotel_table.configure(yscroll=scrollbar.set)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            self.hotel_table.pack(fill=tk.BOTH, expand=True)

            # Обработчик выбора отеля
            self.hotel_table.bind('<<TreeviewSelect>>', self.on_hotel_select)

    def create_hotel_details(self):
        """Создание панели деталей отеля."""
        # Создаем фрейм для деталей отеля
        self.details_content = ttk.Frame(self.details_frame)
        self.details_content.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Заголовок с названием отеля
        self.hotel_name_var = tk.StringVar(value="Выберите отель из списка")
        self.hotel_name_label = ttk.Label(self.details_content,
                                          textvariable=self.hotel_name_var,
                                          font=("Arial", 14, "bold"),
                                          wraplength=300)
        self.hotel_name_label.pack(pady=(0, 10))

        # Рейтинг отеля (звезды)
        self.rating_frame = ttk.Frame(self.details_content)
        self.rating_frame.pack(fill=tk.X, pady=(0, 10))

        self.rating_label = ttk.Label(self.rating_frame, text="Рейтинг: ")
        self.rating_label.pack(side=tk.LEFT)

        self.stars_frame = ttk.Frame(self.rating_frame)
        self.stars_frame.pack(side=tk.LEFT)

        # Адрес отеля
        self.address_var = tk.StringVar()
        self.address_frame = ttk.Frame(self.details_content)
        self.address_frame.pack(fill=tk.X, pady=(0, 10))

        self.address_label = ttk.Label(self.address_frame, text="Адрес: ")
        self.address_label.pack(side=tk.LEFT, anchor=tk.N)

        self.address_value = ttk.Label(self.address_frame,
                                       textvariable=self.address_var,
                                       wraplength=300)
        self.address_value.pack(side=tk.LEFT)

        # Notebook для отзывов и дополнительной информации
        self.details_notebook = ttk.Notebook(self.details_content)
        self.details_notebook.pack(fill=tk.BOTH, expand=True, pady=(10, 0))

        # Вкладка с отзывами
        self.reviews_frame = ttk.Frame(self.details_notebook)
        self.details_notebook.add(self.reviews_frame, text="Отзывы")

        # Вкладка с дополнительной информацией (API)
        self.api_info_frame = ttk.Frame(self.details_notebook)
        self.details_notebook.add(self.api_info_frame, text="Информация из API")

        # Настройка вкладки с отзывами
        self.reviews_canvas = tk.Canvas(self.reviews_frame, highlightthickness=0)
        self.reviews_scrollbar = ttk.Scrollbar(self.reviews_frame,
                                               orient="vertical",
                                               command=self.reviews_canvas.yview)
        self.reviews_scrollframe = ttk.Frame(self.reviews_canvas)

        self.reviews_scrollframe.bind(
            "<Configure>",
            lambda e: self.reviews_canvas.configure(
                scrollregion=self.reviews_canvas.bbox("all")
            )
        )

        self.reviews_canvas.create_window((0, 0), window=self.reviews_scrollframe, anchor="nw")
        self.reviews_canvas.configure(yscrollcommand=self.reviews_scrollbar.set)

        self.reviews_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.reviews_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Настройка вкладки с API-информацией
        self.api_info_text = tk.Text(self.api_info_frame, wrap=tk.WORD, width=40, height=20)
        self.api_info_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        api_info_scrollbar = ttk.Scrollbar(self.api_info_frame, command=self.api_info_text.yview)
        api_info_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.api_info_text.config(yscrollcommand=api_info_scrollbar.set)

        # Сообщение о выборе отеля
        self.no_hotel_message = ttk.Label(self.reviews_scrollframe,
                                          text="Выберите отель из списка для просмотра отзывов",
                                          font=("Arial", 10, "italic"))
        self.no_hotel_message.pack(pady=20)

    def refresh_hotel_list(self):
        """Обновление списка отелей из базы данных."""
        try:
            # Получаем список отелей
            self.hotels = self.db.get_hotels()

            # Очищаем таблицу
            for item in self.hotel_table.get_children():
                self.hotel_table.delete(item)

            # Заполняем таблицу данными
            for hotel in self.hotels:
                hotel_id, name, address, review_count, avg_rating = hotel

                # Форматируем рейтинг как количество звезд
                rating_text = f"{avg_rating}★" if avg_rating else "-"

                self.hotel_table.insert('', 'end', values=(
                    hotel_id, name, rating_text, review_count or 0
                ))

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить список отелей: {e}")

    def on_hotel_select(self, event):
        """Обработчик выбора отеля в таблице."""
        # Получаем ID выбранного отеля
        selected_items = self.hotel_table.selection()

        if not selected_items:
            return

        # Получаем данные выбранного отеля
        item = selected_items[0]
        hotel_id = self.hotel_table.item(item, "values")[0]

        # Загружаем информацию об отеле
        self.load_hotel_details(hotel_id)

    def load_hotel_details(self, hotel_id):
        """Загрузка подробной информации об отеле."""
        try:
            # Получаем информацию из базы данных
            hotel, reviews = self.db.get_hotel_details(int(hotel_id))

            if not hotel:
                messagebox.showerror("Ошибка", "Отель не найден в базе данных.")
                return

            # Сохраняем текущий отель
            self.selected_hotel = hotel
            self.hotel_reviews = reviews

            # Обновляем информацию в интерфейсе
            self.update_hotel_info(hotel, reviews)

            # Загружаем дополнительную информацию через API
            self.load_api_info(hotel[1])  # hotel[1] - название отеля

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить информацию об отеле: {e}")

    def update_hotel_info(self, hotel, reviews):
        """Обновление информации об отеле в интерфейсе."""
        # Обновляем название отеля
        self.hotel_name_var.set(hotel[1])  # hotel[1] - название отеля

        # Обновляем адрес
        self.address_var.set(hotel[2] or "Адрес не указан")  # hotel[2] - адрес

        # Обновляем рейтинг (звезды)
        for widget in self.stars_frame.winfo_children():
            widget.destroy()

        avg_rating = hotel[4]  # hotel[4] - средний рейтинг
        if avg_rating:
            star_count = int(avg_rating)

            for i in range(5):
                if i < star_count:
                    star_text = "★"  # Закрашенная звезда
                    star_color = "#FFD700"  # Золотой цвет
                else:
                    star_text = "☆"  # Пустая звезда
                    star_color = "#C0C0C0"  # Серый цвет

                star_label = ttk.Label(self.stars_frame, text=star_text,
                                       font=("Arial", 16), foreground=star_color)
                star_label.pack(side=tk.LEFT)
        else:
            ttk.Label(self.stars_frame, text="Нет оценок",
                      font=("Arial", 10, "italic")).pack(side=tk.LEFT)

        # Обновляем отзывы
        # Очищаем фрейм с отзывами
        for widget in self.reviews_scrollframe.winfo_children():
            widget.destroy()

        if not reviews:
            ttk.Label(self.reviews_scrollframe, text="Нет отзывов",
                      font=("Arial", 10, "italic")).pack(pady=20)
            return

        # Добавляем отзывы
        for i, review_data in enumerate(reviews):
            review = review_data['review']
            criteria = review_data['criteria_ratings']

            review_frame = ttk.Frame(self.reviews_scrollframe)
            review_frame.pack(fill=tk.X, pady=(0, 15), padx=5)

            # Разделяем отзывы линией
            if i > 0:
                separator = ttk.Separator(review_frame, orient='horizontal')
                separator.pack(fill=tk.X, pady=(0, 10))

            # Заголовок отзыва
            header_frame = ttk.Frame(review_frame)
            header_frame.pack(fill=tk.X, pady=(5, 10))

            # Рейтинг отзыва
            rating = review[1]  # review[1] - рейтинг
            rating_label = ttk.Label(header_frame, text=f"Рейтинг: {rating}★",
                                     font=("Arial", 10, "bold"))
            rating_label.pack(side=tk.LEFT)

            # Дата отзыва
            date = review[3]  # review[3] - дата
            if date:
                date_parts = date.split(" ")[0].split("-")
                formatted_date = f"{date_parts[2]}.{date_parts[1]}.{date_parts[0]}"
                date_label = ttk.Label(header_frame, text=f"Дата: {formatted_date}",
                                       font=("Arial", 10))
                date_label.pack(side=tk.RIGHT)

            # Группируем критерии по категориям
            categories = {}
            for criteria_rating in criteria:
                category = criteria_rating[3]  # criteria_rating[3] - название категории

                if category not in categories:
                    categories[category] = []

                categories[category].append(criteria_rating)

            # Отображаем критерии по категориям
            for category, criteria_list in categories.items():
                category_frame = ttk.LabelFrame(review_frame, text=category)
                category_frame.pack(fill=tk.X, pady=(0, 5))

                # Создаем таблицу для критериев
                criteria_table = ttk.Treeview(category_frame, columns=('criteria', 'rating', 'description'),
                                              show='headings', height=len(criteria_list))

                criteria_table.heading('criteria', text='Критерий')
                criteria_table.heading('rating', text='Оценка')
                criteria_table.heading('description', text='Описание')

                criteria_table.column('criteria', width=150)
                criteria_table.column('rating', width=70, anchor=tk.CENTER)
                criteria_table.column('description', width=200)

                for criteria_rating in criteria_list:
                    name = criteria_rating[0]  # criteria_rating[0] - название критерия
                    rating = criteria_rating[1]  # criteria_rating[1] - оценка
                    description = criteria_rating[2]  # criteria_rating[2] - описание

                    criteria_table.insert('', 'end', values=(name, rating, description))

                criteria_table.pack(fill=tk.X, expand=True)

    def load_api_info(self, hotel_name):
        """Загрузка информации об отеле через API."""
        try:
            # Очищаем текстовое поле
            self.api_info_text.config(state=tk.NORMAL)
            self.api_info_text.delete(1.0, tk.END)

            # Получаем информацию из Booking.com
            booking_info = self.api.search_hotels(hotel_name)

            if not booking_info:
                self.api_info_text.insert(tk.END,
                                          "Не удалось найти информацию об отеле через API.\n\n" +
                                          "Попробуйте использовать кнопку 'Поиск через API' " +
                                          "для ручного поиска.")
                self.api_info_text.config(state=tk.DISABLED)
                return

            # Форматируем информацию
            self.api_info_text.insert(tk.END, "ИНФОРМАЦИЯ ИЗ ГОСТИНИЧНЫХ СЕРВИСОВ\n\n", "header")

            self.api_info_text.insert(tk.END, "Найденные отели:\n\n", "subheader")

            for hotel in booking_info[:3]:  # Показываем только первые 3 результата
                self.api_info_text.insert(tk.END, f"Название: {hotel['name']}\n", "name")
                self.api_info_text.insert(tk.END, f"Адрес: {hotel['address']}\n\n")

            # Получаем детальную информацию о первом отеле
            if booking_info:
                hotel_details = self.api.get_hotel_details(booking_info[0]['id'])

                self.api_info_text.insert(tk.END, "ДЕТАЛЬНАЯ ИНФОРМАЦИЯ\n\n", "header")

                if hotel_details:
                    self.api_info_text.insert(tk.END, f"Название: {hotel_details.get('name', '')}\n", "name")
                    self.api_info_text.insert(tk.END, f"Описание: {hotel_details.get('description', '')}\n")
                    self.api_info_text.insert(tk.END, f"Адрес: {hotel_details.get('address', '')}\n")
                    self.api_info_text.insert(tk.END, f"Рейтинг: {hotel_details.get('rating', 0)}★\n\n")

                    # Удобства
                    if hotel_details.get('amenities'):
                        self.api_info_text.insert(tk.END, "Удобства:\n", "subheader")
                        for amenity in hotel_details['amenities'][:10]:  # Первые 10 удобств
                            self.api_info_text.insert(tk.END, f"• {amenity}\n")

                        if len(hotel_details['amenities']) > 10:
                            self.api_info_text.insert(tk.END,
                                                      f"... и еще {len(hotel_details['amenities']) - 10} удобств\n")

                        self.api_info_text.insert(tk.END, "\n")

                # Получаем отзывы
                hotel_reviews = self.api.get_hotel_reviews(booking_info[0]['id'], 3)

                if hotel_reviews:
                    self.api_info_text.insert(tk.END, "ОТЗЫВЫ\n\n", "header")

                    for review in hotel_reviews:
                        self.api_info_text.insert(tk.END, f"Автор: {review['author']}\n", "name")
                        self.api_info_text.insert(tk.END, f"Дата: {review['date']}\n")
                        self.api_info_text.insert(tk.END, f"Оценка: {review['rating']}/10\n")
                        self.api_info_text.insert(tk.END, f"Заголовок: {review['title']}\n")
                        self.api_info_text.insert(tk.END, f"Отзыв: {review['text']}\n")

                        if review.get('positive'):
                            self.api_info_text.insert(tk.END, f"Плюсы: {review['positive']}\n")

                        if review.get('negative'):
                            self.api_info_text.insert(tk.END, f"Минусы: {review['negative']}\n")

                        self.api_info_text.insert(tk.END, "\n")

            # Настраиваем стили текста
            self.api_info_text.tag_configure("header", font=("Arial", 12, "bold"))
            self.api_info_text.tag_configure("subheader", font=("Arial", 10, "bold"))
            self.api_info_text.tag_configure("name", font=("Arial", 10, "bold"))

            self.api_info_text.config(state=tk.DISABLED)

        except Exception as e:
            self.api_info_text.delete(1.0, tk.END)
            self.api_info_text.insert(tk.END, f"Ошибка при загрузке информации через API: {e}")
            self.api_info_text.config(state=tk.DISABLED)

    def search_hotels(self):
        """Поиск отелей в базе данных."""
        search_query = self.search_var.get().strip().lower()

        if not search_query:
            self.refresh_hotel_list()
            return

        try:
            # Очищаем таблицу
            for item in self.hotel_table.get_children():
                self.hotel_table.delete(item)

            # Фильтруем отели по запросу
            filtered_hotels = [
                hotel for hotel in self.hotels
                if search_query in hotel[1].lower()  # hotel[1] - название отеля
            ]

            # Заполняем таблицу отфильтрованными данными
            for hotel in filtered_hotels:
                hotel_id, name, address, review_count, avg_rating = hotel

                # Форматируем рейтинг как количество звезд
                rating_text = f"{avg_rating}★" if avg_rating else "-"

                self.hotel_table.insert('', 'end', values=(
                    hotel_id, name, rating_text, review_count or 0
                ))

            # Обновляем статус
            hotel_count = len(filtered_hotels)
            if hotel_count > 0:
                messagebox.showinfo("Результаты поиска", f"Найдено отелей: {hotel_count}")
            else:
                messagebox.showinfo("Результаты поиска", "Отели не найдены")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при поиске отелей: {e}")

    def reset_search(self):
        """Сброс результатов поиска."""
        self.search_var.set("")
        self.refresh_hotel_list()

    def search_hotel_api(self):
        """Поиск отеля через API."""
        from tkinter import simpledialog

        # Запрашиваем название отеля
        hotel_name = simpledialog.askstring("Поиск отеля",
                                            "Введите название отеля для поиска:",
                                            parent=self)

        if not hotel_name:
            return

        try:
            # Создаем новое окно для результатов
            api_window = tk.Toplevel(self)
            api_window.title(f"Поиск отеля: {hotel_name}")
            api_window.geometry("800x600")

            # Добавляем рамку
            api_frame = ttk.Frame(api_window, padding=10)
            api_frame.pack(fill=tk.BOTH, expand=True)

            # Заголовок
            ttk.Label(api_frame, text=f"Результаты поиска для: {hotel_name}",
                      font=("Arial", 14, "bold")).pack(pady=(0, 10))

            # Выполняем поиск
            hotels = self.api.search_hotels(hotel_name)

            if not hotels:
                ttk.Label(api_frame, text="Отели не найдены или API ключ не настроен",
                          font=("Arial", 12)).pack(pady=20)
                ttk.Label(api_frame, text="Для использования API получите ключ на сайте Яндекс.Путешествий:",
                          font=("Arial", 10)).pack(pady=5)
                ttk.Label(api_frame, text="https://yandex.ru/dev/travel-partners-api/",
                          font=("Arial", 10, "underline"), foreground="blue").pack(pady=5)
                return

            # Создаем таблицу для отелей
            columns = ('name', 'address', 'stars', 'rating')
            hotels_table = ttk.Treeview(api_frame, columns=columns, show='headings', height=10)

            hotels_table.heading('name', text='Название')
            hotels_table.heading('address', text='Адрес')
            hotels_table.heading('stars', text='Звезды')
            hotels_table.heading('rating', text='Рейтинг')

            hotels_table.column('name', width=250)
            hotels_table.column('address', width=300)
            hotels_table.column('stars', width=80, anchor=tk.CENTER)
            hotels_table.column('rating', width=80, anchor=tk.CENTER)

            # Заполняем таблицу результатами
            for hotel in hotels:
                hotels_table.insert('', 'end', values=(
                    hotel['name'],
                    hotel['address'],
                    f"{hotel.get('stars', 0)}★",
                    hotel.get('rating', 0)
                ))

            # Добавляем полосу прокрутки
            scrollbar = ttk.Scrollbar(api_frame, orient=tk.VERTICAL, command=hotels_table.yview)
            hotels_table.configure(yscroll=scrollbar.set)

            # Размещаем таблицу и полосу прокрутки
            hotels_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, pady=10)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=10)

            # Фрейм для деталей отеля
            details_frame = ttk.LabelFrame(api_frame, text="Детали отеля")
            details_frame.pack(fill=tk.BOTH, expand=True, pady=10)

            # Изначально показываем сообщение о выборе
            ttk.Label(details_frame, text="Выберите отель из списка для просмотра деталей",
                      font=("Arial", 10, "italic")).pack(pady=20)

            # Обработчик выбора отеля
            def on_hotel_select(event):
                # Очищаем фрейм деталей
                for widget in details_frame.winfo_children():
                    widget.destroy()

                # Получаем выбранный отель
                selection = hotels_table.selection()
                if not selection:
                    return

                # Получаем индекс выбранного отеля
                index = hotels_table.index(selection[0])
                hotel = hotels[index]

                # Получаем детали отеля
                hotel_details = self.api.get_hotel_details(hotel['id'])

                if not hotel_details:
                    ttk.Label(details_frame, text="Не удалось получить информацию об отеле",
                              font=("Arial", 10, "italic")).pack(pady=20)
                    return

                # Название отеля
                ttk.Label(details_frame, text=hotel_details.get('name', ""),
                          font=("Arial", 12, "bold")).pack(anchor=tk.W, pady=(10, 5))

                # Адрес
                ttk.Label(details_frame, text=f"Адрес: {hotel_details.get('address', '')}",
                          wraplength=700).pack(anchor=tk.W)

                # Рейтинг
                rating_frame = ttk.Frame(details_frame)
                rating_frame.pack(anchor=tk.W, pady=5)

                ttk.Label(rating_frame,
                          text=f"Рейтинг: {hotel_details.get('rating', 0)} | Звезды: {hotel_details.get('stars', 0)}★").pack(
                    side=tk.LEFT)

                # Описание
                if hotel_details.get('description'):
                    desc_frame = ttk.LabelFrame(details_frame, text="Описание")
                    desc_frame.pack(fill=tk.X, pady=5)

                    ttk.Label(desc_frame, text=hotel_details.get('description', ""),
                              wraplength=700).pack(padx=5, pady=5)

                # Удобства
                if hotel_details.get('amenities'):
                    amenities_frame = ttk.LabelFrame(details_frame, text="Удобства")
                    amenities_frame.pack(fill=tk.X, pady=5)

                    # Размещаем удобства в 3 колонки
                    amenities_grid = ttk.Frame(amenities_frame)
                    amenities_grid.pack(fill=tk.X, padx=5, pady=5)

                    for i, amenity in enumerate(hotel_details.get('amenities', [])[:12]):
                        row = i // 3
                        col = i % 3
                        ttk.Label(amenities_grid, text=f"• {amenity}").grid(row=row, column=col, sticky=tk.W, padx=5,
                                                                            pady=2)

                # Кнопка добавления отеля
                ttk.Button(details_frame, text="Добавить отель в базу данных",
                           command=lambda: self.add_api_hotel_to_db(hotel['name'],
                                                                    address=hotel_details.get('address', ''))).pack(
                    pady=10)

            # Привязываем обработчик к таблице
            hotels_table.bind('<<TreeviewSelect>>', on_hotel_select)

        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при поиске отеля: {e}")

    def show_hotel_details(self, hotel):
        """Отображение подробной информации об отеле."""
        # Создаем новое окно для деталей
        details_window = tk.Toplevel(self)
        details_window.title(f"Детали отеля: {hotel['name']}")
        details_window.geometry("700x500")

        # Добавляем рамку
        details_frame = ttk.Frame(details_window, padding=10)
        details_frame.pack(fill=tk.BOTH, expand=True)

        # Заголовок
        ttk.Label(details_frame, text=hotel['name'],
                  font=("Arial", 16, "bold")).pack(pady=(0, 10))

        # Загружаем детальную информацию
        hotel_details = self.api.get_hotel_details(hotel['id'])

        if not hotel_details:
            ttk.Label(details_frame, text="Детальная информация не найдена",
                      font=("Arial", 10, "italic")).pack(pady=20)
            return

        # Основная информация
        info_frame = ttk.Frame(details_frame)
        info_frame.pack(fill=tk.X, pady=10)

        # Адрес
        ttk.Label(info_frame, text=f"Адрес: {hotel_details.get('address', 'Не указан')}",
                  wraplength=650).pack(anchor=tk.W, pady=2)

        # Рейтинг
        ttk.Label(info_frame, text=f"Рейтинг: {hotel_details.get('rating', 'Не указан')} | " +
                                   f"Звёзды: {hotel_details.get('stars', 'Не указаны')}★").pack(anchor=tk.W, pady=2)

        # Контактная информация
        if hotel_details.get('phone') or hotel_details.get('email') or hotel_details.get('website'):
            contact_frame = ttk.LabelFrame(details_frame, text="Контактная информация")
            contact_frame.pack(fill=tk.X, pady=5)

            if hotel_details.get('phone'):
                ttk.Label(contact_frame, text=f"Телефон: {hotel_details.get('phone')}").pack(anchor=tk.W, padx=5,
                                                                                             pady=2)

            if hotel_details.get('email'):
                ttk.Label(contact_frame, text=f"Email: {hotel_details.get('email')}").pack(anchor=tk.W, padx=5, pady=2)

            if hotel_details.get('website'):
                website_frame = ttk.Frame(contact_frame)
                website_frame.pack(anchor=tk.W, padx=5, pady=2)

                ttk.Label(website_frame, text=f"Сайт: {hotel_details.get('website')}").pack(side=tk.LEFT)

                ttk.Button(website_frame, text="Открыть",
                           command=lambda: webbrowser.open(hotel_details.get('website'))).pack(side=tk.LEFT, padx=5)

        # Описание
        if hotel_details.get('description'):
            desc_frame = ttk.LabelFrame(details_frame, text="Описание")
            desc_frame.pack(fill=tk.X, pady=5)

            desc_text = tk.Text(desc_frame, wrap=tk.WORD, height=5)
            desc_text.insert(tk.END, hotel_details.get('description'))
            desc_text.config(state=tk.DISABLED)
            desc_text.pack(fill=tk.X, padx=5, pady=5)

        # Время заезда/выезда
        if hotel_details.get('check_in_time') or hotel_details.get('check_out_time'):
            checkin_frame = ttk.Frame(details_frame)
            checkin_frame.pack(fill=tk.X, pady=5)

            if hotel_details.get('check_in_time'):
                ttk.Label(checkin_frame, text=f"Время заезда: {hotel_details.get('check_in_time')}").pack(anchor=tk.W)

            if hotel_details.get('check_out_time'):
                ttk.Label(checkin_frame, text=f"Время выезда: {hotel_details.get('check_out_time')}").pack(anchor=tk.W)

        # Удобства
        if hotel_details.get('amenities'):
            amenities_frame = ttk.LabelFrame(details_frame, text="Удобства")
            amenities_frame.pack(fill=tk.X, pady=5)

            # Создаем сетку для удобств (по 3 в ряд)
            amenities_grid = ttk.Frame(amenities_frame)
            amenities_grid.pack(fill=tk.X, padx=5, pady=5)

            for i, amenity in enumerate(hotel_details.get('amenities')):
                row = i // 3
                col = i % 3
                ttk.Label(amenities_grid, text=f"• {amenity}").grid(row=row, column=col, sticky=tk.W, padx=5, pady=2)

        # Кнопка добавления в базу данных
        ttk.Button(details_frame, text="Добавить отель в базу данных",
                   command=lambda: self.add_api_hotel_to_db(hotel['name'],
                                                            address=hotel_details.get('address', ''))).pack(pady=10)

    def add_api_hotel_to_db(self, hotel_name, address=""):
        """Добавление отеля из API в базу данных."""
        try:
            # Добавляем отель в базу данных
            hotel_id = self.db.add_hotel(hotel_name, address)

            messagebox.showinfo("Успешно",
                                f"Отель '{hotel_name}' успешно добавлен в базу данных.")

            # Обновляем список отелей
            self.refresh_hotel_list()

            # Предлагаем перейти к оценке отеля
            if messagebox.askyesno("Оценка отеля",
                                   f"Хотите оценить отель '{hotel_name}' прямо сейчас?"):
                if hasattr(self.parent.master, 'new_rating'):
                    self.parent.master.new_rating()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось добавить отель в базу данных: {e}")

    def export_hotel_list(self):
        """Экспорт списка отелей в файл."""
        from tkinter import filedialog
        import csv

        # Запрашиваем имя файла для сохранения
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить список отелей"
        )

        if not file_path:
            return

        try:
            # Если выбран формат CSV
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)

                    # Записываем заголовки
                    writer.writerow(['ID', 'Название', 'Адрес', 'Количество отзывов', 'Средний рейтинг'])

                    # Записываем данные отелей
                    for hotel in self.hotels:
                        writer.writerow(hotel)

            # Если выбран формат Excel
            elif file_path.lower().endswith('.xlsx'):
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment

                    # Создаем новую книгу и лист
                    workbook = openpyxl.Workbook()
                    sheet = workbook.active
                    sheet.title = "Список отелей"

                    # Записываем заголовки
                    headers = ['ID', 'Название', 'Адрес', 'Количество отзывов', 'Средний рейтинг']
                    for col_num, header in enumerate(headers, 1):
                        cell = sheet.cell(row=1, column=col_num)
                        cell.value = header
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center')

                    # Записываем данные отелей
                    for row_num, hotel in enumerate(self.hotels, 2):
                        for col_num, value in enumerate(hotel, 1):
                            cell = sheet.cell(row=row_num, column=col_num)
                            cell.value = value

                    # Автоматическая настройка ширины столбцов
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter

                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))

                        adjusted_width = (max_length + 2) * 1.2
                        sheet.column_dimensions[column_letter].width = adjusted_width

                    # Сохраняем книгу
                    workbook.save(file_path)

                except ImportError:
                    messagebox.showwarning("Предупреждение",
                                           "Для экспорта в Excel требуется модуль openpyxl. " +
                                           "Будет создан файл CSV.")

                    # Если модуль не установлен, сохраняем в CSV
                    csv_path = file_path.replace('.xlsx', '.csv')

                    with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
                        writer = csv.writer(csvfile)

                        # Записываем заголовки
                        writer.writerow(['ID', 'Название', 'Адрес', 'Количество отзывов', 'Средний рейтинг'])

                        # Записываем данные отелей
                        for hotel in self.hotels:
                            writer.writerow(hotel)

                    file_path = csv_path

            messagebox.showinfo("Экспорт завершен",
                                f"Список отелей успешно экспортирован в файл:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать список отелей: {e}")